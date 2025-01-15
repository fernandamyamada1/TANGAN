import tensorflow as tf
import numpy as np
from tensorflow.keras.layers import LeakyReLU
from tensorflow.keras.models import load_model
from numpy import ones
from numpy import zeros
from tensorflow.keras.initializers import RandomNormal
import cv2 as cv
from openpyxl import Workbook, load_workbook
import os
import gc
from tensorflow.keras.layers import Input, Dense, BatchNormalization, Conv2D, Concatenate
from tensorflow.keras.layers import Dropout, GlobalAveragePooling2D
import logging
from tensorflow.compat.v1 import ConfigProto
from tensorflow.compat.v1 import InteractiveSession

config = ConfigProto()
config.gpu_options.allow_growth = True
session = InteractiveSession(config=config)

train = True
retrain = False

tf.keras.backend.clear_session()
tf.config.run_functions_eagerly(True)

batch = 2
tones = [31, 255, 62, 155, 124, 93, 186]
image_shape = 512

logging.basicConfig(filename='gan_train.log',
                    filemode='w',
                    encoding='utf-8',
                    format='%(asctime)s %(levelname)s:%(message)s',
                    datefmt='%m/%d/%Y %I:%M:%S %p',
                    level=logging.INFO)

console = logging.StreamHandler()
console.setLevel(logging.INFO)

logging.getLogger().addHandler(console)
logger = logging.getLogger('gan_train')

# gpus = tf.config.experimental.list_physical_devices('GPU')
# if len(gpus) > 0:
#     print("USING GPU")
#     tf.config.experimental.set_virtual_device_configuration(gpus[0], [tf.config.experimental.VirtualDeviceConfiguration(memory_limit=10000)])

def save_model(i, disc_model1, gen_model1, gan_model1):
    filename = './model/disc1.h5'
    disc_model1.save_weights(filename)
    filename = './model/gen1.h5'
    gen_model1.save_weights(filename)
    filename = './model/gan1.h5'
    gan_model1.save_weights(filename)
    file = open("./model/checkpoint.txt", "w")
    file.write(str(i))
    file.close()
    return

def wmae(y_true, y_pred, penalty_factor=5):
    absolute_diff = tf.abs(y_true - y_pred)
    penalty = tf.where(tf.equal(y_true, 0), penalty_factor * absolute_diff, absolute_diff)
    return tf.reduce_mean(penalty)

def load_model(disc_model1, gen_model1, gan_model1):
    filename = './model/disc1.h5'
    disc_model1.load_weights(filename)
    filename = './model/gen1.h5'
    gen_model1.load_weights(filename)
    filename = './model/gan1.h5'
    gan_model1.load_weights(filename)
    return disc_model1, gen_model1, gan_model1

def calculate_hu_moments(grayscale_image):
    # Convert the grayscale image to float32
    grayscale_image = tf.cast(grayscale_image, dtype=tf.float32)

    # Normalize the grayscale image
    mean = tf.reduce_mean(grayscale_image)
    std = tf.math.reduce_std(grayscale_image)
    normalized_image = (grayscale_image - mean) / (std + 1e-10)  # Add a small value to avoid division by zero

    # Calculate the image moments
    x = tf.range(tf.shape(normalized_image)[1], dtype=tf.float32)
    y = tf.range(tf.shape(normalized_image)[0], dtype=tf.float32)
    x, y = tf.meshgrid(x, y)
    x = tf.reshape(x, [-1])
    y = tf.reshape(y, [-1])
    flattened_image = tf.reshape(normalized_image, [-1])
    m00 = tf.reduce_sum(flattened_image)
    m10 = tf.reduce_sum(x * flattened_image)
    m01 = tf.reduce_sum(y * flattened_image)
    m20 = tf.reduce_sum(x ** 2 * flattened_image)
    m02 = tf.reduce_sum(y ** 2 * flattened_image)
    m11 = tf.reduce_sum(x * y * flattened_image)
    m30 = tf.reduce_sum(x ** 3 * flattened_image)
    m03 = tf.reduce_sum(y ** 3 * flattened_image)
    m12 = tf.reduce_sum(x * y ** 2 * flattened_image)
    m21 = tf.reduce_sum(x ** 2 * y * flattened_image)

    # Calculate the Hu Moments
    hu1 = m20 + m02
    hu2 = (m20 - m02) ** 2 + 4 * m11 ** 2
    hu3 = (m30 - 3 * m12) ** 2 + (3 * m21 - m03) ** 2
    hu4 = (m30 + m12) ** 2 + (m21 + m03) ** 2
    hu5 = (m30 - 3 * m12) * (m30 + m12) * ((m30 + m12) ** 2 - 3 * (m21 + m03) ** 2) + \
          (3 * m21 - m03) * (m21 + m03) * (3 * (m30 + m12) ** 2 - (m21 + m03) ** 2)
    hu6 = (m20 - m02) * ((m30 + m12) ** 2 - (m21 + m03) ** 2) + \
          4 * m11 * (m30 + m12) * (m21 + m03)

    # Return the Hu Moments as a tensor
    hu_moments = tf.stack([hu1, hu2, hu3, hu4, hu5, hu6])
    return hu_moments

def tensor_threshold(x, tensor, t = 0.05):
    # Calculate lower and upper bounds
    lower_bound = x - t
    upper_bound = x + t

    # Create boolean masks for values within the threshold range
    mask_lower = tf.greater_equal(tensor, lower_bound)
    mask_upper = tf.less_equal(tensor, upper_bound)

    # Combine the masks using logical AND
    combined_mask = tf.logical_and(mask_lower, mask_upper)

    # Convert boolean mask to integers (True -> 1, False -> 0)
    combined_mask = tf.cast(combined_mask, tf.int32)

    return combined_mask
    
def hu_loss(y_true, y_pred):
  
    y_true= tf.divide(y_true, 255.0)
    y_pred = tf.divide(y_pred, 255.0)

    sum = tf.constant(0, dtype=tf.float32)
    data = tf.stack([y_true, y_pred], axis=1)

    for i in data:
        pred = i[0]
        true = i[1]

        oc_pred = tf.cast(tf.math.not_equal(pred, tf.constant(0, dtype=tf.float32)), tf.float32)

        oc_true = tf.cast(tf.math.not_equal(true, tf.constant(0, dtype=tf.float32)), tf.float32)


        for tone in tones:
            tone = tf.constant(tone/255, dtype=tf.float32)

            mask_pred = tensor_threshold(tone, pred)

            n_pred = tf.reduce_sum(mask_pred)

            mask_true = tensor_threshold(tone, true)

            n_true = tf.reduce_sum(mask_true)

            hu_pred = calculate_hu_moments(mask_pred)

            hu_true = calculate_hu_moments(mask_true)

            n_true = tf.cast(n_true, tf.float32)

            n_pred = tf.cast(n_pred, tf.float32)

            diff = tf.math.abs(tf.math.subtract(n_pred, n_true))

            ratio = diff / (diff + 1) 

            ratio = tf.cast(ratio, tf.float32)

            hu_true = tf.where(tf.math.is_nan(hu_true), tf.zeros_like(hu_true), hu_true)

            hu_true = tf.where(tf.math.is_inf(hu_true), tf.zeros_like(hu_true), hu_true)

            hu_pred = tf.where(tf.math.is_nan(hu_pred), tf.zeros_like(hu_pred), hu_pred)

            hu_pred = tf.where(tf.math.is_inf(hu_pred), tf.zeros_like(hu_pred), hu_pred)

            dist = tf.abs(tf.subtract(hu_true, hu_pred))

            dist = tf.cast(dist, tf.float32)

            d1 = dist + tf.ones_like(dist)

            dist = dist / d1

            dist = tf.reduce_sum(dist) * ratio

            sum = tf.math.add(sum, tf.math.divide(dist * 0.5, (len(tones) * batch * 6)))
    return sum

def discriminator(image_shape):
    init = RandomNormal(stddev=0.02)
    dropout_rate = 0.4
    in_src_image = Input(shape=image_shape)
    in_target_image = Input(shape=image_shape)
    merged = Concatenate()([in_src_image, in_target_image])

    d = Conv2D(64, (4, 4), strides=(2, 2), padding='same', kernel_initializer=init)(merged)
    d = LeakyReLU(alpha=0.2)(d)
    d = Conv2D(128, (4, 4), strides=(2, 2), padding='same', kernel_initializer=init)(d)
    d = BatchNormalization()(d)
    d = LeakyReLU(alpha=0.2)(d)

    d = Conv2D(256, (4, 4), strides=(2, 2), padding='same', kernel_initializer=init)(d)
    d = BatchNormalization()(d)
    d = LeakyReLU(alpha=0.2)(d)
    d = Dropout(dropout_rate)(d)

    d = Conv2D(512, (4, 4), strides=(2, 2), padding='same', kernel_initializer=init)(d)
    d = BatchNormalization()(d)
    d = LeakyReLU(alpha=0.2)(d)
    d = Dropout(dropout_rate)(d)

    d = Conv2D(1024, (4, 4), padding='same', kernel_initializer=init)(d)
    d = BatchNormalization()(d)
    d = LeakyReLU(alpha=0.2)(d)
    d = Dropout(dropout_rate)(d)

    d = GlobalAveragePooling2D()(d)
    d = Dense(1, activation='sigmoid')(d)

    model = tf.keras.models.Model([in_src_image, in_target_image], d)
    opt = tf.keras.optimizers.Adam(learning_rate=0.0002, beta_1=0.5)
    loss = tf.keras.losses.BinaryCrossentropy()
    model.compile(loss=loss, optimizer=opt)

    return model

class Sampling(tf.keras.layers.Layer):
    def call(self, inputs):
        z_mean, z_log_var = inputs
        batch = tf.shape(z_mean)[0]
        dim = tf.shape(z_mean)[1]
        epsilon = tf.random.normal(shape=(batch, dim, 1, 2))
        return z_mean + tf.exp(0.5 * z_log_var) * epsilon

def generator(input_shape):

    latent_dim = 2

    input_layer = tf.keras.layers.Input(shape=input_shape)
    x1 = tf.keras.layers.Conv2D(48, kernel_size=(11, 11), strides=(1,1), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(input_layer)
    x1 = tf.keras.layers.BatchNormalization(axis=-1)(x1)
    x1 = tf.keras.layers.LeakyReLU(alpha=0.2)(x1)

    x1_0 = tf.keras.layers.Conv2D(48, kernel_size=(9, 9), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(x1)
    x1_0 = tf.keras.layers.Dropout(.2)(x1_0)
    x1_0 = tf.keras.layers.BatchNormalization(axis=-1)(x1_0)
    x1_0 = tf.keras.layers.LeakyReLU(alpha=0.2)(x1_0)

    x2 = tf.keras.layers.Conv2D(48, kernel_size=(9, 9), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(x1_0)
    x2 = tf.keras.layers.Dropout(.2)(x2)
    x2 = tf.keras.layers.BatchNormalization(axis=-1)(x2)
    x2 = tf.keras.layers.LeakyReLU(alpha=0.2)(x2)

    x2_0 = tf.keras.layers.Conv2D(48, kernel_size=(9, 9), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(x2)
    x2_0 = tf.keras.layers.Dropout(.2)(x2_0)
    x2_0 = tf.keras.layers.BatchNormalization(axis=-1)(x2_0)
    x2_0 = tf.keras.layers.LeakyReLU(alpha=0.2)(x2_0)

    x3 = tf.keras.layers.Conv2D(48, kernel_size=(7, 7), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(x2_0)
    x3 = tf.keras.layers.Dropout(.2)(x3)
    x3 = tf.keras.layers.BatchNormalization(axis=-1)(x3)
    x3 = tf.keras.layers.LeakyReLU(alpha=0.2)(x3)

    x3_0 = tf.keras.layers.Conv2D(48, kernel_size=(7, 7), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(x3)
    x3_0 = tf.keras.layers.Dropout(.2)(x3_0)
    x3_0 = tf.keras.layers.BatchNormalization(axis=-1)(x3_0)
    x3_0 = tf.keras.layers.LeakyReLU(alpha=0.2)(x3_0)

    x4 = tf.keras.layers.Conv2D(48, kernel_size=(5, 5), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(x3_0)
    x4 = tf.keras.layers.Dropout(.2)(x4)
    x4 = tf.keras.layers.BatchNormalization(axis=-1)(x4)
    x4 = tf.keras.layers.LeakyReLU(alpha=0.2)(x4)

    x4_0 = tf.keras.layers.Conv2D(48, kernel_size=(5, 5), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(x4)
    x4_0 = tf.keras.layers.Dropout(.2)(x4_0)
    x4_0 = tf.keras.layers.BatchNormalization(axis=-1)(x4_0)
    x4_0 = tf.keras.layers.LeakyReLU(alpha=0.2)(x4_0)

    x5 = tf.keras.layers.Conv2D(48, kernel_size=(3, 3), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(x4_0)
    x5 = tf.keras.layers.BatchNormalization(axis=-1)(x5)
    x5 = tf.keras.layers.LeakyReLU(alpha=0.2)(x5)

    z_mean = tf.keras.layers.Dense(latent_dim, name="z_mean")(x5)
    z_log_var = tf.keras.layers.Dense(latent_dim, name="z_log_var")(x5)
    z = Sampling()([z_mean, z_log_var])


    tr1 = tf.keras.layers.Conv2DTranspose(48, kernel_size=(5, 5), strides=(2, 2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(z)
    tr1 = tf.keras.layers.BatchNormalization(axis=-1)(tr1)
    tr1 = tf.keras.layers.LeakyReLU(alpha=0.2)(tr1)

    tr2 = tf.keras.layers.Conv2DTranspose(48, kernel_size=(7, 7), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(tf.keras.layers.Concatenate()([tr1, x4_0]))
    tr2 = tf.keras.layers.Dropout(.2)(tr2)
    tr2 = tf.keras.layers.BatchNormalization(axis=-1)(tr2)
    tr2 = tf.keras.layers.LeakyReLU(alpha=0.2)(tr2)

    tr2_0 = tf.keras.layers.Conv2DTranspose(48, kernel_size=(7, 7), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(tf.keras.layers.Concatenate()([tr2, x4]))
    tr2_0 = tf.keras.layers.Dropout(.2)(tr2_0)
    tr2_0 = tf.keras.layers.BatchNormalization(axis=-1)(tr2_0)
    tr2_0 = tf.keras.layers.LeakyReLU(alpha=0.2)(tr2_0)

    tr3 = tf.keras.layers.Conv2DTranspose(48, kernel_size=(9, 9), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(tf.keras.layers.Concatenate()([tr2_0, x3_0]))
    tr3 = tf.keras.layers.Dropout(.2)(tr3)
    tr3 = tf.keras.layers.BatchNormalization(axis=-1)(tr3)
    tr3 = tf.keras.layers.LeakyReLU(alpha=0.2)(tr3)

    tr3_0 = tf.keras.layers.Conv2DTranspose(48, kernel_size=(9, 9), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(tf.keras.layers.Concatenate()([tr3, x3]))
    tr3_0 = tf.keras.layers.Dropout(.2)(tr3_0)
    tr3_0 = tf.keras.layers.BatchNormalization(axis=-1)(tr3_0)
    tr3_0 = tf.keras.layers.LeakyReLU(alpha=0.2)(tr3_0)

    tr4 = tf.keras.layers.Conv2DTranspose(48, kernel_size=(11, 11), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(tf.keras.layers.Concatenate()([tr3_0, x2_0]))
    tr4 = tf.keras.layers.Dropout(.2)(tr4)
    tr4 = tf.keras.layers.BatchNormalization(axis=-1)(tr4)
    tr4 = tf.keras.layers.LeakyReLU(alpha=0.2)(tr4)

    tr4_0 = tf.keras.layers.Conv2DTranspose(48, kernel_size=(11, 11), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(tf.keras.layers.Concatenate()([tr4, x2]))
    tr4_0 = tf.keras.layers.Dropout(.2)(tr4_0)
    tr4_0 = tf.keras.layers.BatchNormalization(axis=-1)(tr4_0)
    tr4_0 = tf.keras.layers.LeakyReLU(alpha=0.2)(tr4_0)


    tr5 = tf.keras.layers.Conv2DTranspose(48, kernel_size=(11, 11), strides=(2,2), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(tf.keras.layers.Concatenate()([tr4_0, x1_0]))
    tr5 = tf.keras.layers.Dropout(.2)(tr5)
    tr5 = tf.keras.layers.BatchNormalization(axis=-1)(tr5)
    tr5 = tf.keras.layers.LeakyReLU(alpha=0.2)(tr5)

    output = tf.keras.layers.Conv2D(input_shape[2], kernel_size=(1, 1), strides=(1,1), padding='same', kernel_regularizer=tf.keras.regularizers.l2())(tf.keras.layers.Concatenate()([tr5, x1]))
    output = tf.keras.layers.BatchNormalization(axis=-1)(output)
    output = tf.keras.layers.Activation("softplus")(output)

    model = tf.keras.Model(inputs=[input_layer], outputs=[output])
    model.compile()

    return model

def create_model(gen1, disc1, input_img):
    opt = tf.keras.optimizers.legacy.Adam(learning_rate=0.0002, beta_1=0.5)
    disc1.trainable = False
    src_input1 = tf.keras.layers.Input(shape=input_img)
    gen_output1 = gen1(src_input1)
    disc_output1 = disc1([src_input1, gen_output1])
    gan1 = tf.keras.models.Model(inputs=src_input1, outputs=[disc_output1, gen_output1])
    gan1.compile(loss=["binary_crossentropy", hu_loss], optimizer=opt, loss_weights=[1, 100])
    return gan1

def generate_fake_images(model, sample, patch_shape):
    X = model.predict(sample)
    y = zeros((len(X), patch_shape, patch_shape, 1))
    return X,y

def approx(tensor):
    size = tensor.shape[0]
    out = []

    # Transform into tensor
    approximation_tensor = tf.constant([0] + tones, dtype=tf.float32)

    # Reshape the approximation tensor to (1, 1, 1, 8) to match the input tensor shape
    approximation_tensor_reshaped = tf.reshape(approximation_tensor, (1, 1, 1, 8))

    for i in range(size):
        input_tensor = tensor[i]

        # Compute the absolute difference between the input tensor and the approximation tensor
        absolute_difference = tf.abs(input_tensor - approximation_tensor_reshaped)

        # Find the index of the minimum absolute difference along the last dimension
        min_index = tf.argmin(absolute_difference, axis=-1)

        # Use the index to approximate the input tensor values
        approximated_tensor = tf.gather(approximation_tensor, min_index)

        # Ensure the output tensor has the same shape as the input tensor
        approximated_tensor = tf.reshape(approximated_tensor, input_tensor.shape)

        out.append(approximated_tensor)

    tensor = tf.stack(out)

    return tensor


def performance_check(step, testA_generator, testB_generator, disc_model1, gen_model1, gan_model1):
    total_inferences = tf.data.experimental.cardinality(testA_generator).numpy().item() 
    flag = "test"

    i = 0

    for (realA, _), (realB, _) in zip(testA_generator, testB_generator):

        fakeB, _ = generate_fake_images(gen_model1, realA, batch)

        idx = (i) * batch

        ids = testA_generator.file_paths[idx: idx + testA_generator._batch_size]

        ids = [os.path.basename(file_path) for file_path in ids]

        fakeB = approx(fakeB)

        score_gen = gen_model1.evaluate(fakeB, realB, verbose=0)

        path = './train/' + str(step+1) + flag + '/'

        if not os.path.exists(path):
            os.makedirs(path)

        save_images(fakeB, path, ids)

        i = i + 1

        if i >= total_inferences:
            break

    save_model(step, disc_model1, gen_model1, gan_model1)

    logger.info(f"Score: {score_gen}")

def save_images(tensor, path, ids):
    size = tensor.shape[0]

    for i in range(size):
        img = tensor[i]
        pil_img = tf.keras.preprocessing.image.array_to_img(img)

        opencv_image = np.array(pil_img)

        # Save the image using OpenCV
        cv.imwrite(path + str(ids[i]), opencv_image)

def augment(tensor, random_rotation, random_flip_h, random_flip_v, random_zoom):
    size = tensor.shape[0]
    out = []
    for i in range(size):
        img= tensor[i]
        img= tf.image.rot90(img, k = random_rotation[i])
        if random_flip_h[i] == 1:
            img = tf.image.flip_left_right(img)
        if random_flip_v[i] == 1:
            img = tf.image.flip_up_down(img)
        img = tf.image.resize_with_pad(img, random_zoom[i], random_zoom[i])
        img = tf.image.resize(img, (tensor.shape[1], tensor.shape[2]))
        out.append(img)
    tensor = tf.stack(out)
    return tensor

def train_gan(disc_model, gen_model, gan_model, path, epochs=600, val_interval= 100, test_interval=100):

    global tones
    
    loss = np.zeros(3)
    gc.enable()

    trainA_generator = tf.keras.utils.image_dataset_from_directory(
        path + 'trvalA/',  # Path to trainA directory
        image_size=(image_shape, image_shape),
        batch_size=batch,
        color_mode='grayscale',
        seed=42,  # Use the same seed for both generators to maintain correspondence
    )

    trainB_generator = tf.keras.utils.image_dataset_from_directory(
        path + 'trvalB/',  # Path to trainA directory
        image_size=(image_shape, image_shape),
        batch_size=batch,
        color_mode='grayscale',
        seed=42,  # Use the same seed for both generators to maintain correspondence
    )

    testA_generator = tf.keras.utils.image_dataset_from_directory(
        path + 'testA/',  # Path to trainA directory
        image_size=(image_shape, image_shape),
        batch_size=batch,
        color_mode='grayscale',
        seed=42,  # Use the same seed for both generators to maintain correspondence
    )

    testB_generator = tf.keras.utils.image_dataset_from_directory(
        path + 'testB/',  # Path to trainA directory
        image_size=(image_shape, image_shape),
        batch_size=batch,
        color_mode='grayscale',
        seed=42,  # Use the same seed for both generators to maintain correspondence
    )

    if not retrain:
        path = "./loss.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "GEN1"
        sheet["B1"] = "DISC1_1"
        sheet["C1"] = "DISC1_2"

        workbook.save(path)

    total_steps = tf.data.experimental.cardinality(trainA_generator).numpy().item() 
    print("Total steps: ", str(total_steps))

    ini = 0

    if retrain:
        with open("./model/checkpoint.txt") as file:
                ini = int(file.read()) + 1
    
    for epoch in range(ini, epochs):
        logger.info("\nStart of epoch %d" % (epoch,))
        print("GAN Step: ", epoch, "/", epochs)        
        disc_model, gen_model, gan_model, loss = run_epoch(disc_model, gen_model, gan_model, loss, trainA_generator, trainB_generator, total_steps)
        print("END EPOCH!")
        print("Step: ", epoch)
        path = "./loss.xlsx"
        workbook = load_workbook(filename=path)
        sheet = workbook.active
        avg_loss = [str(round(l, 3) / val_interval) for l in loss]
        print(">%d, g[%.3f] d1[%.3f] d2[%.3f]" % ((epoch + 1), float(avg_loss[0]), float(avg_loss[1]), float(avg_loss[2])))
        sheet.append(avg_loss)
        workbook.save(path)
        loss = np.zeros(3)
        if (epoch + 1) % val_interval == 0:
            performance_check(epoch, testA_generator, testB_generator, disc_model, gen_model, gan_model)

def run_epoch(disc_model, gen_model, gan_model, loss, trainA_generator, trainB_generator, total_steps):
    data = zip(trainA_generator, trainB_generator)
    iterations = 0
    n_patch = disc_model.output_shape[1]
    
    train_loss = []
    val_loss = []
    gen_loss = []
    
    for (realA, _), (realB, _) in data:
        print("Iteration: ", iterations)
        iterations = iterations + 1

        y_real = ones((batch, 1))
        y_fake = zeros((batch, 1))

        realA = add_gaussian_noise(realA)

        random_rotation = np.random.randint(4, size=batch)
        random_flip_v = np.random.randint(2, size=batch)
        random_flip_h = np.random.randint(2, size=batch)
        random_zoom = np.random.uniform(low=1.00 * image_shape, high=1.10 * image_shape, size=batch).astype(int)

        realA = augment(realA, random_rotation, random_flip_h, random_flip_v, random_zoom)
        realB = augment(realB, random_rotation, random_flip_h, random_flip_v, random_zoom)

        fakeB, _ = generate_fake_images(gen_model, realA, n_patch)
        fakeB = approx(fakeB)

        disc_loss1_1 = disc_model.train_on_batch([realA, realB], y_real)
        disc_loss1_2 = disc_model.train_on_batch([realA, fakeB], y_fake)
        gen_loss1, _, _ = gan_model.train_on_batch(realA, [y_real, realB])

        loss[0] += gen_loss1
        loss[1] += disc_loss1_1
        loss[2] += disc_loss1_2

        train_loss.append(disc_loss1_1)
        val_loss.append(disc_loss1_2)
        gen_loss.append(gen_loss1)

        if iterations >= total_steps:
            break
    logger.info(f"Train Loss (Mean):  {sum(train_loss) / len(train_loss)}")
    logger.info(f"Val Loss (Mean): {sum(val_loss) / len(val_loss)}")
    logger.info(f"Gen Loss (Mean): {sum(gen_loss) / len(gen_loss)}")
    tf.keras.backend.clear_session()
    return disc_model, gen_model, gan_model, loss

def add_gaussian_noise(image):
    std = 0.1
    noise = tf.random.normal(shape=tf.shape(image), mean=0.0, stddev=std, dtype=tf.float32) 
    return image + noise

def main():
    path = "./dataset/"

    shape = (image_shape, image_shape, 1)

    disc_model = discriminator(shape)
    gen_model = generator(shape)
    gan_model = create_model(gen_model, disc_model, shape)

    if retrain:
        disc_model, gen_model, gan_model = load_model(disc_model, gen_model, gan_model)

    train_gan(disc_model, gen_model, gan_model, path)

if __name__ == "__main__":
    main()
